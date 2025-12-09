# -*- coding: utf-8 -*-
"""
teklif_app.py
Excel (XLSX) çıktı üreten Teklif Uygulaması
- Malzeme tablosu (satır ekle/sil, inline edit)
- Ödeme planı tablosu (satır ekle/sil, inline edit)
- Excel çıktısı A4'e göre hizalanmış, firma bilgileri, müşteri kutusu, toplamlar, imza alanı
- Logo üstte ortalı (varsa ef.png / logo.png kullanılır)
Requires: openpyxl, pillow (optional for image embedding)
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import sys
from pathlib import Path
import json
import webbrowser
import tempfile
import subprocess

# Excel imports (must be installed)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Optional: image embedding in Excel (requires Pillow)
try:
    from openpyxl.drawing.image import Image as XLImage
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

class TeklifApp:
    def __init__(self, root):
        self.root = root
        self.root.title("EF Yapı Dekorasyon - Teklif Uygulaması (Excel)")
        try:
            self.root.state('zoomed')
        except Exception:
            pass
        self.root.configure(bg='#f0f0f0')

        # Ayarlar dosyası
        self.config_file = Path(__file__).parent /"setup"/ "teklif_ayarlari.json"
        self.settings = self.load_settings()

        # Uygulama dosyasının bulunduğu klasör
        app_dir = os.path.dirname(os.path.abspath(__file__))

        # Logo: aranan dosyalar
        self.logo_path = None
        for lf in ("ef.png","logo.png"):
            p = os.path.join(app_dir, lf)
            if os.path.exists(p):
                self.logo_path = p
                break

        # Ana frame
        main_frame = tk.Frame(root, bg='#f0f0f0', padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Başlık (UI başlığı)
        title_frame = tk.Frame(main_frame, bg='#f0f0f0')
        title_frame.pack(fill=tk.X, pady=(0, 10))

        title_label = tk.Label(title_frame, text="EF YAPI DEKORASYON",
                               font=('Arial', 26, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title_label.pack()

        subtitle_label = tk.Label(title_frame, text="Fiyat Teklif Uygulaması (Excel Çıktısı)",
                                  font=('Arial', 18), bg='#f0f0f0', fg='#7f8c8d')
        subtitle_label.pack()

        # Müşteri bilgileri frame
        customer_frame = tk.LabelFrame(main_frame, text="Müşteri Bilgileri",
                                       font=('Arial', 10, 'bold'), bg='#f0f0f0', padx=10, pady=10)
        customer_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(customer_frame, text="Ad Soyad:", bg='#f0f0f0', font=('Arial', 9)).grid(row=0, column=0, sticky='w', padx=6, pady=4)
        self.customer_name = tk.Entry(customer_frame, width=30, font=('Arial', 9))
        self.customer_name.grid(row=0, column=1, padx=6, pady=4)

        tk.Label(customer_frame, text="T.C. Kimlik No:", bg='#f0f0f0', font=('Arial', 9)).grid(row=0, column=2, sticky='w', padx=6, pady=4)
        self.customer_tc = tk.Entry(customer_frame, width=25, font=('Arial', 9))
        self.customer_tc.grid(row=0, column=3, padx=6, pady=4)

        tk.Label(customer_frame, text="Telefon:", bg='#f0f0f0', font=('Arial', 9)).grid(row=1, column=0, sticky='w', padx=6, pady=4)
        self.customer_phone = tk.Entry(customer_frame, width=30, font=('Arial', 9))
        self.customer_phone.grid(row=1, column=1, padx=6, pady=4)

        tk.Label(customer_frame, text="Adres:", bg='#f0f0f0', font=('Arial', 9)).grid(row=2, column=0, sticky='w', padx=6, pady=4)
        self.customer_address = tk.Entry(customer_frame, width=60, font=('Arial', 9))
        self.customer_address.grid(row=2, column=1, columnspan=3, padx=6, pady=4)


        # Menü çubuğu
        self.menubar = tk.Menu(root)
        root.config(menu=self.menubar)
        self.create_menu()

        # Butonlar frame
        button_frame = tk.Frame(main_frame, bg='#f0f0f0')
        button_frame.pack(fill=tk.X, pady=(0, 6))

        preview_btn = tk.Button(button_frame, text="Excel Önizleme", command=self.preview_excel,
                                bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                                padx=18, pady=5, cursor='hand2')
        preview_btn.pack(side=tk.LEFT, padx=5)

        save_btn = tk.Button(button_frame, text="Excel Kaydet (Yazdırma için Excel'den yazdırın)", command=self.save_excel,
                              bg='#27ae60', fg='white', font=('Arial', 10, 'bold'),
                              padx=18, pady=5, cursor='hand2')
        save_btn.pack(side=tk.LEFT, padx=5)

        # Tablolar container
        tables_container = tk.Frame(main_frame, bg='#f0f0f0')
        tables_container.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Sol taraf - Malzeme tablosu
        material_container = tk.LabelFrame(tables_container, text="MALZEME TABLOSU",
                                           font=('Arial', 11, 'bold'), bg='#f0f0f0', padx=5, pady=5)
        material_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        self.setup_table(material_container, "material")

        # Sağ taraf - Ödeme planı
        payment_container = tk.LabelFrame(tables_container, text="ÖDEME PLANI",
                                          font=('Arial', 11, 'bold'), bg='#f0f0f0', padx=5, pady=5)
        payment_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        self.setup_table(payment_container, "payment")

        # Genel toplam frame
        total_frame = tk.LabelFrame(main_frame, text="Genel Toplam", font=('Arial', 10, 'bold'), bg='#f0f0f0', padx=10, pady=10)
        total_frame.pack(fill=tk.X)
        self.total_without_vat_label = tk.Label(total_frame, text="KDV Hariç Toplam: 0.00 ₺",
                                                font=('Arial', 11, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        self.total_without_vat_label.pack(side=tk.LEFT, padx=20)
        self.total_with_vat_label = tk.Label(total_frame, text="KDV Dahil Toplam: 0.00 ₺",
                                             font=('Arial', 11, 'bold'), bg='#f0f0f0', fg='#27ae60')
        self.total_with_vat_label.pack(side=tk.LEFT, padx=20)

    # ---------------- Settings ----------------
    def load_settings(self):
        """Ayarları yükle"""
        default_settings = {'save_folder': None}
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    s = json.load(f)
                    default_settings.update(s)
        except Exception:
            pass
        return default_settings

    def save_settings(self):
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.settings, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Hata", f"Ayarlar kaydedilirken hata oluştu:\n{e}")

    def get_save_folder(self):
        if self.settings.get('save_folder'):
            return Path(self.settings['save_folder'])
        else:
            return Path(__file__).parent / "Teklifler"

    def get_save_folder_display(self):
        folder = self.get_save_folder()
        try:
            return str(folder)
        except:
            return "Varsayılan (Uygulama Klasörü)"

    def select_save_folder(self):
        current_folder = self.get_save_folder()
        folder = filedialog.askdirectory(title="Tekliflerin Kaydedileceği Klasörü Seçin",
                                         initialdir=str(current_folder) if current_folder.exists() else None)
        if folder:
            self.settings['save_folder'] = str(Path(folder))
            self.save_settings()
            messagebox.showinfo("Başarılı", f"Kayıt klasörü güncellendi:\n{folder}")
            self.root.after(100, self.update_menu)

    def reset_save_folder(self):
        self.settings['save_folder'] = None
        self.save_settings()
        messagebox.showinfo("Başarılı", "Varsayılan klasöre dönüldü.")
        self.root.after(100, self.update_menu)

    def create_menu(self):
        try:
            self.menubar.delete(0, tk.END)
        except Exception:
            pass
        ayarlar_menu = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="Ayarlar", menu=ayarlar_menu)
        ayarlar_menu.add_command(label="Kayıt Klasörü Seç", command=self.select_save_folder)
        ayarlar_menu.add_command(label="Varsayılan Klasöre Dön", command=self.reset_save_folder)
        ayarlar_menu.add_separator()
        current_folder = self.get_save_folder_display()
        if len(current_folder) > 50:
            current_folder = "..." + current_folder[-47:]
        ayarlar_menu.add_command(label=f"Mevcut: {current_folder}", state='disabled')

    def update_menu(self):
        self.create_menu()

    # ----------------- Table UI -----------------
    def setup_table(self, parent, table_type):
        """Tablo yapısını oluştur (material veya payment)"""
        table_frame = tk.Frame(parent, bg='#f0f0f0')
        table_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        scrollbar = tk.Scrollbar(table_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        if table_type == "material":
            columns = ('Ürün/İşçilik Adı', 'Birim', 'Miktar', 'Birim Fiyat', 'Toplam')
        else:
            columns = ('Tarih', 'Genel Toplam', 'Alınacak Tutar', 'Kalan Tutar')

        tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=12, yscrollcommand=scrollbar.set)
        scrollbar.config(command=tree.yview)

        for col in columns:
            tree.heading(col, text=col)
        # column widths
        if table_type == "material":
            tree.column('Ürün/İşçilik Adı', width=220, anchor='w')
            tree.column('Birim', width=70, anchor='center')
            tree.column('Miktar', width=70, anchor='center')
            tree.column('Birim Fiyat', width=100, anchor='e')
            tree.column('Toplam', width=100, anchor='e')
        else:
            tree.column('Tarih', width=90, anchor='center')
            tree.column('Genel Toplam', width=110, anchor='e')
            tree.column('Alınacak Tutar', width=110, anchor='e')
            tree.column('Kalan Tutar', width=110, anchor='e')

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Buttons
        btn_frame = tk.Frame(parent, bg='#f0f0f0')
        btn_frame.pack(fill=tk.X, padx=6, pady=6)
        if table_type == "material":
            add_btn = tk.Button(btn_frame, text="Satır Ekle", command=lambda: self.add_row(tree, table_type),
                                bg='#3498db', fg='white', font=('Arial', 9), padx=8, pady=3)
            add_btn.pack(side=tk.LEFT, padx=5)
            delete_btn = tk.Button(btn_frame, text="Satır Sil", command=lambda: self.delete_row(tree, table_type),
                                   bg='#e74c3c', fg='white', font=('Arial', 9), padx=8, pady=3)
            delete_btn.pack(side=tk.LEFT, padx=5)
        else:
            add_btn = tk.Button(btn_frame, text="Ödeme Satırı Ekle", command=lambda: self.add_row(tree, table_type),
                                bg='#3498db', fg='white', font=('Arial', 9), padx=8, pady=3)
            add_btn.pack(side=tk.LEFT, padx=5)
            delete_btn = tk.Button(btn_frame, text="Satır Sil", command=lambda: self.delete_row(tree, table_type),
                                   bg='#e74c3c', fg='white', font=('Arial', 9), padx=8, pady=3)
            delete_btn.pack(side=tk.LEFT, padx=5)

        # Totals area for material
        if table_type == "material":
            total_frame = tk.Frame(parent, bg='#f0f0f0')
            total_frame.pack(fill=tk.X, padx=6, pady=(4,6))
            tk.Label(total_frame, text="Ara Toplam:", bg='#f0f0f0', font=('Arial', 9)).grid(row=0, column=0, sticky='w', padx=3)
            subtotal_value = tk.Label(total_frame, text="0.00 ₺", bg='#f0f0f0', font=('Arial', 9))
            subtotal_value.grid(row=0, column=1, sticky='e', padx=3)
            tk.Label(total_frame, text="KDV (%20):", bg='#f0f0f0', font=('Arial', 9)).grid(row=0, column=2, sticky='w', padx=30)
            vat_value = tk.Label(total_frame, text="0.00 ₺", bg='#f0f0f0', font=('Arial', 9))
            vat_value.grid(row=0, column=3, sticky='e', padx=3)
            self.material_subtotal = subtotal_value
            self.material_vat = vat_value

        # Save tree refs
        if table_type == "material":
            self.material_tree = tree
        else:
            self.payment_tree = tree

        # Inline editing support (simple)
        tree.bind('<Double-1>', lambda e: self.start_edit(tree, table_type, e))
        tree.bind('<Button-1>', lambda e: self.on_click(tree, e))
        tree.bind('<Return>', lambda e: self.finish_edit(tree, table_type))
        tree.bind('<Escape>', lambda e: self.cancel_edit(tree))
        tree.bind('<Delete>', lambda e: self.delete_row(tree, table_type))
        tree.bind('<KeyRelease>', lambda e: self.update_totals())

    def add_row(self, tree, table_type):
        if table_type == "material":
            item = tree.insert('', 'end', values=('', 'Adet', '1', '0.00', '0.00'))
        else:
            # payment default columns: Tarih, Genel Toplam, Alınacak, Kalan
            item = tree.insert('', 'end', values=(datetime.now().strftime('%d.%m.%Y'), '0.00', '0.00', '0.00'))
        self.update_totals()

    def delete_row(self, tree, table_type):
        sel = tree.selection()
        if sel:
            for s in sel:
                tree.delete(s)
            self.update_totals()
        else:
            messagebox.showwarning("Uyarı", "Lütfen silmek için bir satır seçin.")

    def start_edit(self, tree, table_type, event):
        # End previous edit
        if hasattr(self, 'editing_cells') and tree in self.editing_cells:
            self.finish_edit(tree, table_type)

        region = tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        col_index = int(column.replace('#','')) - 1
        if not item:
            return
        bbox = tree.bbox(item, column)
        if not bbox:
            return
        x, y, width, height = bbox
        values = list(tree.item(item, 'values'))
        current_value = values[col_index] if col_index < len(values) else ''
        entry = tk.Entry(tree, font=('Arial', 9))
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, str(current_value))
        entry.select_range(0, tk.END)
        entry.focus()
        if not hasattr(self, 'editing_cells'):
            self.editing_cells = {}
        self.editing_cells[tree] = {'entry': entry, 'item': item, 'col_index': col_index, 'table_type': table_type}
        entry.bind('<Return>', lambda e: self.finish_edit(tree, table_type))
        entry.bind('<Escape>', lambda e: self.cancel_edit(tree))

    def on_click(self, tree, event):
        if hasattr(self, 'editing_cells') and tree in self.editing_cells:
            self.finish_edit(tree, self.editing_cells[tree]['table_type'])

    def finish_edit(self, tree, table_type):
        if not hasattr(self, 'editing_cells') or tree not in self.editing_cells:
            return
        info = self.editing_cells[tree]
        entry = info['entry']
        item = info['item']
        col_index = info['col_index']
        new_value = entry.get()
        entry.destroy()
        values = list(tree.item(item, 'values'))
        # ensure list length
        if col_index >= len(values):
            values += [''] * (col_index - len(values) + 1)
        values[col_index] = new_value
        # If numeric fields, try to sanitize
        if tree == getattr(self, 'material_tree', None):
            if col_index in (2,3):  # Miktar veya Birim Fiyat changed
                try:
                    miktar = float(values[2]) if values[2] else 0.0
                    birim = float(values[3]) if values[3] else 0.0
                    values[4] = f"{miktar * birim:.2f}"
                except Exception:
                    values[4] = "0.00"
        else:
    # Payment table otomatik Kalan Tutar ve alt satır ekleme
            try:
             # Alınacak Tutar mevcut ve 0'dan farklıysa işlem yap
                alinacak_str = values[2] if len(values) > 2 else ''
                alinacak = float(alinacak_str) if alinacak_str.strip() != '' else 0.0
                if alinacak != 0.0:
                    genel_toplam = float(values[1]) if len(values) > 1 and values[1] else 0.0
                    kalan = genel_toplam - alinacak

                    if kalan < 0:
                        from tkinter import messagebox
                        messagebox.showwarning("Uyarı", "Alınacak tutar Genel Toplamdan büyük olamaz!")
                        kalan = 0.0  # kalan sıfır olarak kalır
                    if len(values) < 4:
                        values += ['0.00'] * (4 - len(values))
                    values[3] = f"{kalan:.2f}"

            # Eğer kalan > 0 ve bu son satırsa alt satır ekle
                    last_item = tree.get_children()[-1] if tree.get_children() else None
                    if kalan > 0 and item == last_item:
                        tree.insert('', 'end', values=(
                            datetime.now().strftime('%d.%m.%Y'),
                            f"{kalan:.2f}",
                            "0.00",
                            f"{kalan:.2f}"
                        ))
            except Exception as e:
                print("Payment row update error:", e)



        tree.item(item, values=values)
        del self.editing_cells[tree]
        self.update_totals()

    def cancel_edit(self, tree):
        if hasattr(self, 'editing_cells') and tree in self.editing_cells:
            entry = self.editing_cells[tree]['entry']
            try:
                entry.destroy()
            except:
                pass
            del self.editing_cells[tree]

    def calculate_table_totals(self, tree):
        subtotal = 0.0
        for item in tree.get_children():
            values = tree.item(item, 'values')
            try:
                # material: total at index 4
                if tree == getattr(self, 'material_tree', None):
                    total = float(str(values[4]).replace('₺','').strip()) if len(values) > 4 and values[4] else 0.0
                else:
                    # payment not included in totals
                    total = 0.0
                subtotal += total
            except Exception:
                pass
        without_vat = subtotal
        vat_amount = round(without_vat * 0.20, 2)
        with_vat = round(without_vat + vat_amount, 2)
        return round(subtotal,2), round(without_vat,2), round(vat_amount,2), round(with_vat,2)

    def update_totals(self):
        try:
            m_sub, m_without, m_vat, m_with = self.calculate_table_totals(self.material_tree)
            self.material_subtotal.config(text=f"{m_sub:.2f} ₺")
            self.material_vat.config(text=f"{m_vat:.2f} ₺")
            self.total_without_vat_label.config(text=f"KDV Hariç Toplam: {m_without:.2f} ₺")
            self.total_with_vat_label.config(text=f"KDV Dahil Toplam: {m_with:.2f} ₺")
        except Exception:
            pass

    def get_table_data(self, tree):
        data = []
        for item in tree.get_children():
            values = tree.item(item, 'values')
            data.append(values)
        return data

    # ---------------- EXCEL Generation ----------------
    def create_excel(self, output_path):
        """
        Creates a styled Excel file resembling the provided layout with:
        - Ürün/İşçilik Adı ve Müşteri Adres hücrelerinde wrap text (30 karakterde alt satır)
        - Malzeme ve ödeme planı tabloları
        - KDV ve toplam hesaplamaları
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        import tempfile
        from openpyxl.drawing.image import Image as XLImage

        def wrap_text_by_n(text, n=45):
            """Metni her n karakterde alt satıra geçir"""
            if not text:
                return ""
            lines = []
            while len(text) > n:
                split_at = text.rfind(" ", 0, n)
                if split_at == -1:
                    split_at = n
                lines.append(text[:split_at])
                text = text[split_at:].lstrip()
            lines.append(text)
            return "\n".join(lines)

        wb = Workbook()
        ws = wb.active
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 6     # Sıra No
        ws.column_dimensions['B'].width = 45    # Açıklama (geniş)
        ws.column_dimensions['C'].width = 10    # Birim (dar)
        ws.column_dimensions['D'].width = 12    # Miktar
        ws.column_dimensions['E'].width = 14    # Birim fiyat
        ws.column_dimensions['F'].width = 14    # Toplam

        ws.title = "Teklif"

        # Styles
        bold = Font(bold=True)
        big_bold = Font(bold=True, size=14)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        thin = Side(border_style="medium", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="2F4F4F")  # dark grey
        header_fg = "FFFFFF"
        alt_fill = PatternFill("solid", fgColor="ECF0F1")  # light grey

        # Column widths
        col_widths = {1:5, 2:35, 3:12, 4:10, 5:14, 6:14}
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        r = 1

        # Logo
        if self.logo_path and PIL_AVAILABLE:
            try:
                img = XLImage(self.logo_path)
                img.width = 120  # yaklaşık 2 cm
                img.height = 120
                ws.add_image(img, "F1")  # üst sağ veya uygun konum
            except Exception:
                pass

        # Başlık
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="EF YAPI DEKORASYON").font = big_bold
        ws.cell(row=r, column=1).alignment = center
        r += 1

        # Firma info
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="Firma: EF Yapı     ").font = Font(size=14)
        ws.cell(row=r, column=1).alignment = left
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="Firma Sahibi: Fatih AYDIN      Tel: 0537 517 41 19  ").font = Font(size=14)
        ws.cell(row=r, column=1).alignment = left
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="Adres: İbnisina Mahallesi Serkan Sokak No:5/1    E-mail: efyapi0@gmail.com      ").font = Font(size=14)
        ws.cell(row=r, column=1).alignment = left
        r += 2

        # Teklif başlığı
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value="FİYAT TEKLİFİ").font = Font(bold=True, size=12)
        ws.cell(row=r, column=1).alignment = center
        r += 1

        # Müşteri kutusu
        # Müşteri kutusu (gri background)
        box_rows = [
            ("MÜŞTERİ ADI :", self.customer_name.get() if hasattr(self, 'customer_name') else ''),
            ("T.C. :", self.customer_tc.get() if hasattr(self, 'customer_tc') else ''),
            ("TEL :", self.customer_phone.get() if hasattr(self, 'customer_phone') else ''),
            ("ADRES :", self.customer_address.get() if hasattr(self, 'customer_address') else ''),
            ("TARİH :", datetime.now().strftime('%d.%m.%Y'))
        ]

        # Label sütunu = 2 (B), Value = 3..6 (C-F)
        for label, val in box_rows:
            ws.cell(row=r, column=2, value=label).font = Font(bold=True)  # label
            ws.cell(row=r, column=3, value=val)                            # value
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="FFFFFF")
            ws.cell(row=r, column=3).fill = alt_fill
            r += 1

        r += 1

            # Malzeme tablosu başlığı

        
        # ------------------------------------

        headers = ['NO', 'AÇIKLAMA', 'BİRİM', 'MİKTAR', 'BİRİM FİYATI', 'TOPLAM FİYATI']
        # --- SÜTUN GENİŞLİKLERİ (EKLEDİM) ---
       
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True, color=header_fg)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border        

        r += 1
        ws.column_dimensions['A'].width = 6      # NO
        ws.column_dimensions['B'].width = 30     # AÇIKLAMA (GENİŞ)
        ws.column_dimensions['C'].width = 10     # BİRİM (DAR)
        ws.column_dimensions['D'].width = 15     # MİKTAR
        ws.column_dimensions['E'].width = 15     # BİRİM FİYATI
        ws.column_dimensions['F'].width = 15     # TOPLAM FİYATI
        # Malzeme tablosu
        items = self.get_table_data(self.material_tree)
        if not items:
            for i in range(3):
                for c in range(1,7):
                    ws.cell(row=r, column=c, value='').border = border
                r += 1
        else:
            for idx, vals in enumerate(items, start=1):
                name = vals[0] if len(vals) > 0 else ''
                unit = vals[1] if len(vals) > 1 else ''
                qty = vals[2] if len(vals) > 2 else ''
                unit_price = vals[3] if len(vals) > 3 else ''
                total = vals[4] if len(vals) > 4 else ''

                ws.cell(row=r, column=1, value=idx).border = border
                
                ws.cell(row=r, column=2, value=name).border = border
                # Açıklama alt satıra geçsin
                ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                ws.cell(row=r, column=3, value=unit).border = border

                # Miktar
                try:
                    c = ws.cell(row=r, column=4, value=float(qty))
                    c.number_format = '0.00'
                    c.border = border
                except:
                    c = ws.cell(row=r, column=4, value=qty)
                    c.border = border

                # Birim fiyat
                try:
                    c = ws.cell(row=r, column=5, value=float(str(unit_price).replace('₺','').strip()))
                    c.number_format = '#,##0.00'
                    c.border = border
                except:
                    c = ws.cell(row=r, column=5, value=unit_price)
                    c.border = border

                # Toplam fiyat
                try:
                    c = ws.cell(row=r, column=6, value=float(str(total).replace('₺','').strip()))
                    c.number_format = '#,##0.00'
                    c.border = border
                except:
                    c = ws.cell(row=r, column=6, value=total)
                    c.border = border

                # Hizalamalar
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")

                r += 1



        

        # Malzeme toplamları
        m_sub, m_without, m_vat, m_with = self.calculate_table_totals(self.material_tree)
        totals = [("GENEL Toplam:", m_sub), ("KDV'siz Toplam:", m_without),
                ("KDV (%20) Tutarı:", m_vat), ("KDV'li Toplam:", m_with)]
        for label, val in totals:
            ws.cell(row=r, column=5, value=label).font = Font(bold=True)
            ws.cell(row=r, column=6, value=val).number_format = '#,##0.00'
            ws.cell(row=r, column=5).fill = alt_fill
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="FFFFFF")
            ws.cell(row=r, column=5).border = border
            ws.cell(row=r, column=6).border = border
            ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
            r += 1

        r += 1

        # Ödeme planı
        ws.cell(row=r, column=1, value="ÖDEME PLANI").font = Font(bold=True)
        r += 1
        pay_headers = ['NO', 'TARİH', 'TOPLAM', 'ALINACAK TUTAR', 'KALACAK TUTAR','AÇIKLAMALAR']
        for c, h in enumerate(pay_headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True, color=header_fg)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        r += 1
       
        payments = self.get_table_data(self.payment_tree)
        if not payments:
            for i in range(3):
                for c in range(1,6):
                    ws.cell(row=r, column=c, value='').border = border
                r += 1
            # İmza satırı
            ws.cell(row=r+5, column=1, value="İMZA : ________")
            ws.cell(row=r+5, column=5, value="İMZA : ________")
            r += 2  # biraz boşluk bırak
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # yüksekliği sınırsız bırak, sadece genişlik sığsın
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # veya LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4    
        else:
            for idx, vals in enumerate(payments, start=1):
                date = vals[0] if len(vals) > 0 else ''
                full = vals[1] if len(vals) > 1 else ''
                recv = vals[2] if len(vals) > 2 else ''
                rem = vals[3] if len(vals) > 3 else ''

                ws.cell(row=r, column=1, value=idx).alignment = center
                ws.cell(row=r, column=2, value=date).alignment = center
                try:
                    ws.cell(row=r, column=3, value=float(str(full).replace('₺','').strip())).number_format = '#,##0.00'
                except:
                    ws.cell(row=r, column=3, value=full)
                try:
                    ws.cell(row=r, column=4, value=float(str(recv).replace('₺','').strip())).number_format = '#,##0.00'
                except:
                    ws.cell(row=r, column=4, value=recv)
                try:
                    ws.cell(row=r, column=5, value=float(str(rem).replace('₺','').strip())).number_format = '#,##0.00'
                except:
                    ws.cell(row=r, column=5, value=rem)
                for c in range(1,6):
                    ws.cell(row=r, column=c).border = border
                r += 1
                sig_start_row = r+2
                
            # Başlıklar
            ws.cell(row=sig_start_row, column=1, value="MÜŞTERİ").font = Font(bold=True)
            ws.cell(row=sig_start_row, column=5, value="FİRMA YETKİLİSİ").font = Font(bold=True)
            r += 1

            # Tarih satırı
            ws.cell(row=r+3, column=1, value="TARİH : ________")
            ws.cell(row=r+3, column=5, value="TARİH : ________")
            r += 1

            # İmza satırı
            ws.cell(row=r+5, column=1, value="İMZA : ________")
            ws.cell(row=r+5, column=5, value="İMZA : ________")
            r += 2  # biraz boşluk bırak
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # yüksekliği sınırsız bırak, sadece genişlik sığsın
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # veya LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
        # Kaydet
        try:
            wb.save(output_path)
        except Exception:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
            wb.save(tmp)
            import shutil
            shutil.copy(tmp, output_path)
        return output_path
    
    


    def preview_excel(self):
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
            self.create_excel(tmp)
            # open with default app
            webbrowser.open(tmp)
        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulurken hata:\n{e}")

    def save_excel(self):
        customer_name = self.customer_name.get().strip()
        if not customer_name:
            messagebox.showwarning("Uyarı", "Lütfen müşteri adı soyadı girin.")
            return
        try:
            teklifler_dir = self.get_save_folder()
            teklifler_dir.mkdir(parents=True, exist_ok=True)
            # sanitize folder name
            customer_folder_name = customer_name.replace(" ", "_")
            safe_chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_çğıöşüÇĞİÖŞÜ"
            customer_folder_name = ''.join(c if c in safe_chars else '_' for c in customer_folder_name)
            customer_dir = teklifler_dir / customer_folder_name
            customer_dir.mkdir(exist_ok=True)
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Teklif_{date_str}.xlsx"
            path = customer_dir / filename
            self.create_excel(str(path))
            try:
                webbrowser.open(str(path))
            except:
                pass
            messagebox.showinfo("Başarılı", f"Excel oluşturuldu:\n{path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydetme sırasında hata:\n{e}")

if __name__ == "__main__":
    # Locale best-effort (windows tr)
    if sys.platform == 'win32':
        try:
            import locale
            locale.setlocale(locale.LC_ALL, 'Turkish_Turkey.1254')
        except Exception:
            try:
                locale.setlocale(locale.LC_ALL, 'tr_TR.UTF-8')
            except:
                pass
    # Ensure stdout utf-8 if possible
    try:
        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')
        if sys.stderr.encoding != 'utf-8':
            sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

    root = tk.Tk()
    app = TeklifApp(root)
    root.mainloop()
